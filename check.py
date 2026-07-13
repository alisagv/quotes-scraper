from openpyxl.reader.excel import load_workbook

wb = load_workbook("quotes.xlsx")
ws = wb.active

i = 1
for row in ws.iter_rows(values_only=True):
    print(f"{i}.{row}")
    i += 1


